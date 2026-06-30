"""RDTII golden dataset loader — the ground truth for Stage 2 scoring.

Parses the ESCAP RDTII 2.1 Round 1 / Round 2 Excel databases into immutable
``GoldRecord`` rows, and the seed-portal / legal-inventory CSVs into a known-evidence
baseline used by the discovery diff (R20).

Why this replaces the old validation_engine parse: the spreadsheet's ``References``
column *spills* across several unlabelled columns (one URL per column), and the sheets
interleave section-header rows (pillar name, no score) with real indicator rows.
Building the header by dropping ``None`` cells silently shifts every column index and
loses the extra URL columns, so it is done positionally here instead.

Framework-agnostic, no web/LLM imports. Apache-2.0 tooling only (openpyxl: MIT).
"""

from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field

import openpyxl

# RDTII pillars we must cover (R7). Everything else in the workbook is ignored.
MANDATORY_PILLARS = (6, 7)

# A real indicator id looks like "6.1", "7.12". Section-header rows carry a pillar
# *name* in this column instead (e.g. "Cross-border Data Policies"), so this is the
# cheap, reliable way to tell data rows from header rows.
_INDICATOR_RE = re.compile(r"^\d+\.\d+[a-z]?$")
_URL_RE = re.compile(r"^(https?://|www\.)", re.IGNORECASE)

# This module lives at backend/src/zetarix/scoring/; the real workbooks live in the
# repo-root docs/. Four levels up: scoring -> zetarix -> src -> backend -> repo-root.
_DEFAULT_DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", "docs")

_ROUND_FILES = (
    "ESCAP-RDTII-2.1_ Round 1 Database.xlsx",
    "ESCAP-RDTII-2.1_ Round 2 Database.xlsx",
)
_REFERENCE_CSVS = (
    "Sample governemnt portals_Pillar 6_7.csv",
    "Singapore, Malaysia, Australia, Legal Inventory.csv",
)
_IGNORE_SHEETS = {"RDTII 2.1 Methodology", "Consolidated"}


@dataclass(frozen=True)
class GoldRecord:
    """One reviewer-validated RDTII mapping — mirrors domain ``Finding`` fields.

    ``urls`` is the full set of reference links for the mapping (the spilled
    ``References`` columns), normalised order-independent for matching.
    """

    country: str
    pillar_id: int
    indicator_id: str  # golden-DB dotted form, e.g. "6.1"
    act_name: str
    coverage: str  # ↔ Finding.scope
    impact: str  # ↔ Finding.impact
    timeframe: str  # ↔ Finding.last_update (raw string; not parsed to date here)
    raw_score: float | None
    urls: tuple[str, ...] = ()

    @property
    def indicator_canonical(self) -> str:
        """The submission-form code (``6.1`` → ``P6-I1``); additive convenience only."""
        from zetarix.domain.indicator_codes import to_canonical

        return to_canonical(self.indicator_id)


@dataclass(frozen=True)
class ReferenceItem:
    """A known act/portal from the seed CSVs — baseline for the discovery diff (R20)."""

    country: str
    act_name: str
    coverage: str
    timeframe: str
    url: str
    cluster: str = ""


@dataclass(frozen=True)
class _ColumnMap:
    pillar: int
    indicator: int
    score: int
    act: int
    coverage: int
    impact: int
    timeframe: int
    references: int  # URLs span this column to the end of the row


def _clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _detect_columns(row: tuple) -> _ColumnMap | None:
    """Return a positional column map if ``row`` is the header row, else None."""
    idx: dict[str, int] = {}
    for i, cell in enumerate(row):
        name = _clean(cell).lower()
        if not name:
            continue
        if name == "pillar_id":
            idx["pillar"] = i
        elif name == "indicator_id":
            idx["indicator"] = i
        elif name.startswith("raw score"):
            idx["score"] = i
        elif name.startswith("act"):
            idx["act"] = i
        elif name.startswith("coverage"):
            idx["coverage"] = i
        elif name.startswith("impact"):
            idx["impact"] = i
        elif name.startswith("timeframe"):
            idx["timeframe"] = i
        elif name.startswith("reference"):
            idx["references"] = i
    required = ("pillar", "indicator", "act", "coverage", "impact", "timeframe", "references")
    if all(key in idx for key in required):
        return _ColumnMap(**{key: idx[key] for key in (*required, "score")})
    return None


def _collect_urls(row: tuple, start: int) -> tuple[str, ...]:
    """Gather every URL-looking cell from the References column onward."""
    urls: list[str] = []
    for cell in row[start:]:
        text = _clean(cell)
        for part in re.split(r"[\n;]+", text):
            candidate = part.strip()
            if _URL_RE.match(candidate):
                urls.append(candidate)
    return tuple(urls)


def _parse_score(value: object) -> float | None:
    text = _clean(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_sheet(sheet, country: str) -> list[GoldRecord]:
    columns: _ColumnMap | None = None
    records: list[GoldRecord] = []
    for row in sheet.iter_rows(values_only=True):
        if not row or all(cell is None for cell in row):
            continue
        if columns is None:
            columns = _detect_columns(row)
            continue  # header row itself is never data

        pillar_raw = _clean(row[columns.pillar])
        indicator = _clean(row[columns.indicator])
        if not _INDICATOR_RE.match(indicator):
            continue  # section-header row (pillar name, no indicator) — skip
        try:
            pillar_id = int(float(pillar_raw))
        except ValueError:
            continue
        if pillar_id not in MANDATORY_PILLARS:
            continue

        act_name = _clean(row[columns.act])
        urls = _collect_urls(row, columns.references)
        if not act_name and not urls:
            continue  # nothing to match on (blank continuation row) — not ground truth

        records.append(
            GoldRecord(
                country=country,
                pillar_id=pillar_id,
                indicator_id=indicator,
                act_name=act_name,
                coverage=_clean(row[columns.coverage]),
                impact=_clean(row[columns.impact]),
                timeframe=_clean(row[columns.timeframe]),
                raw_score=_parse_score(row[columns.score]),
                urls=urls,
            )
        )
    return records


def load_gold_records(docs_dir: str = _DEFAULT_DOCS_DIR) -> tuple[GoldRecord, ...]:
    """Load all Pillar 6 & 7 mappings from both round databases (ground truth)."""
    records: list[GoldRecord] = []
    for filename in _ROUND_FILES:
        path = os.path.join(docs_dir, filename)
        if not os.path.exists(path):
            continue
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                if sheet_name in _IGNORE_SHEETS:
                    continue
                records.extend(_parse_sheet(workbook[sheet_name], country=sheet_name))
        finally:
            workbook.close()
    return tuple(records)


def load_reference_items(docs_dir: str = _DEFAULT_DOCS_DIR) -> tuple[ReferenceItem, ...]:
    """Load seed-portal / legal-inventory acts — the known-evidence baseline (R20)."""
    items: list[ReferenceItem] = []
    for filename in _REFERENCE_CSVS:
        path = os.path.join(docs_dir, filename)
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row in csv.DictReader(handle):
                url = _clean(row.get("References"))
                items.append(
                    ReferenceItem(
                        country=_clean(row.get("country")),
                        act_name=_clean(row.get("Act.and.or.practice")),
                        coverage=_clean(row.get("Coverage")),
                        timeframe=_clean(row.get("Timeframe")),
                        url=url,
                        cluster=_clean(row.get("cluster")),
                    )
                )
    return tuple(items)
