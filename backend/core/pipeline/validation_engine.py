import openpyxl
import os
import urllib.parse
from typing import List, Dict, Any

class ValidationEngine:
    """Parses RDTII Excel databases to build validation sets for evaluation."""

    def __init__(self, docs_dir: str = "C:/Users/Drew/Desktop/rdtii-autoextract/docs"):
        self.docs_dir = docs_dir
        self.r1_file = os.path.join(docs_dir, "ESCAP-RDTII-2.1_ Round 1 Database.xlsx")
        self.r2_file = os.path.join(docs_dir, "ESCAP-RDTII-2.1_ Round 2 Database.xlsx")

    def load_gold_standard_records(self) -> List[Dict[str, Any]]:
        """Parses R1 and R2 Excel databases and returns consolidated records for Pillars 6 & 7."""
        records = []
        
        # Load Round 1
        if os.path.exists(self.r1_file):
            records.extend(self._parse_workbook(self.r1_file, is_round_1=True))
            
        # Load Round 2
        if os.path.exists(self.r2_file):
            records.extend(self._parse_workbook(self.r2_file, is_round_1=False))
            
        return records

    def _parse_workbook(self, filepath: str, is_round_1: bool) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        records = []
        
        # We ignore helper sheets
        ignore_sheets = ["RDTII 2.1 Methodology", "Consolidated"]
        
        for name in wb.sheetnames:
            if name in ignore_sheets:
                continue
                
            sheet = wb[name]
            country = name
            sheet_records = self._parse_sheet(sheet, country)
            records.extend(sheet_records)
            
        return records

    def _parse_sheet(self, sheet: Any, country: str) -> List[Dict[str, Any]]:
        records = []
        header = None
        
        for row in sheet.iter_rows(values_only=True):
            if not row or all(v is None for v in row):
                continue
                
            # Detect header row (must contain indicator/pillar fields)
            if not header:
                if any(isinstance(val, str) and "Pillar" in val for val in row if val):
                    header = [str(val).strip().lower() for val in row if val is not None]
                continue
                
            # Construct record
            record = {
                "country": country,
                "pillar_id": None,
                "indicator_id": None,
                "act_name": None,
                "coverage": None,
                "impact": None,
                "timeframe": None,
                "references": None
            }
            
            for idx, val in enumerate(row):
                if idx >= len(header):
                    break
                h_name = header[idx]
                
                # Map dynamically based on headers
                if "pillar" in h_name:
                    record["pillar_id"] = val
                elif "indicator" in h_name:
                    record["indicator_id"] = val
                elif "act" in h_name:
                    record["act_name"] = val
                elif "coverage" in h_name:
                    record["coverage"] = val
                elif "impact" in h_name:
                    record["impact"] = val
                elif "timeframe" in h_name:
                    record["timeframe"] = val
                elif "reference" in h_name:
                    record["references"] = val

            # Normalize values
            pillar = record.get("pillar_id")
            if pillar is not None:
                # Convert e.g. 6.0 or "6" or 6 to integer 6
                try:
                    pillar_val = int(float(str(pillar).strip()))
                    record["pillar_id"] = pillar_val
                except ValueError:
                    pass
            
            # Filter specifically for Pillar 6 and 7
            if record["pillar_id"] in (6, 7):
                # Clean up references string (can contain multiple URLs separated by newlines/semicolons)
                ref_str = record.get("references")
                urls = []
                if ref_str:
                    parts = ref_str.replace(";", "\n").split("\n")
                    for p in parts:
                        p_str = p.strip()
                        if p_str.startswith("http"):
                            urls.append(p_str)
                record["parsed_urls"] = urls
                records.append(record)
                
        return records
